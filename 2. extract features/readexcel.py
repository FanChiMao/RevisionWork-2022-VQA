from openpyxl import load_workbook


def read(path, framenum, targetnum):
    wb = load_workbook(path)
    sheetnames = wb.get_sheet_names()
    data = []
    # skip = framenum // targetnum
    frame_index = read2('extract_frame_num_consecutive.xlsx')
    vid = 0
    for sheet in sheetnames:
        count = 1
        count2 = 0
        ws = wb.get_sheet_by_name(sheet)
        # 行迭代
        content = []
        rows = ws.rows
        for row in rows:
            if count2 == targetnum:
                break
            # if count != count2 * skip:
            #     count += 1
            #     continue
            if count != frame_index[0][vid][count2]:
                count += 1
                continue
            line = [col.value for col in row]
            content.append(line)
            count += 1
            count2 += 1
        vid += 1
        data.append(content)  # [video][frame][coefficient]
    return data


def read2(path):
    wb = load_workbook(path)
    sheetnames = wb.get_sheet_names()
    data = []
    for sheet in sheetnames:
        ws = wb.get_sheet_by_name(sheet)
        # 行迭代
        content = []
        rows = ws.rows
        for row in rows:
            line = [col.value for col in row]
            content.append(line)
        data.append(content)  # [video][frame][coefficient]
    return data
